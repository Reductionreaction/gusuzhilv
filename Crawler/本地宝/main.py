import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from urllib.parse import urljoin

base_url = "https://baike.baidu.com"

workbook = Workbook()
sheet = workbook.active

with open('wesites.txt', 'r', encoding='utf-8') as file:
    lines = file.readlines()

# 处理每一行
count=1
for line in lines:
    # 对每一行进行处理，这里以打印为例
    if not line.strip():
        break
    count+=1
    url=line.strip()
    response = requests.get(url)
    response.encoding=response.apparent_encoding
    soup = BeautifulSoup(response.text, 'html.parser')

    # 获取标题

    title = soup.title.string
    leads=soup.find_all('div',class_='lead')
    for lead in leads:
        lead.decompose()
    views = soup.find_all('p', class_='view_city_index')
    for view in views:
        view.decompose()
    ads = soup.find_all('div', class_='adInArticletishi')
    for ad in ads:
        print(ad.text)
        ad.decompose()
    ad_div = soup.find('div', id='add_ewm_content')
    if ad_div:
        ad_div.decompose()
    content=soup.find('div',class_="article-content").text
    print(content)


    print('标题:', title)
    ref1=sheet.cell(row=count, column=1)
    ref2=sheet.cell(row=count, column=2)
    ref1.value=title
    ref2.value=content
workbook.save(filename='牛牛.xlsx')




